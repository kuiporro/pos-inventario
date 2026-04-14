"""
ViewSets del módulo Reportes — corregido para Django 6 + CLP.
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.utils import timezone
from datetime import timedelta

from ventas.models import Venta, VentaDetalle
from inventario.models import Stock, ProductoVariante


class ReporteVentasView(APIView):
    """GET /api/reportes/ventas/?periodo=diario|semanal|mensual"""

    def get(self, request):
        periodo      = request.query_params.get('periodo', 'diario')
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin    = request.query_params.get('fecha_fin')

        qs = Venta.objects.filter(estado='COMPLETADA')

        if fecha_inicio:
            qs = qs.filter(fecha__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha__date__lte=fecha_fin)
        elif not fecha_inicio:
            qs = qs.filter(fecha__gte=timezone.now() - timedelta(days=30))

        trunc_fn = {'mensual': TruncMonth, 'semanal': TruncWeek}.get(periodo, TruncDay)

        try:
            resumen = list(
                qs.annotate(periodo=trunc_fn('fecha'))
                .values('periodo')
                .annotate(
                    cantidad_ventas=Count('id'),
                    total_ventas=Sum('total'),
                )
                .order_by('periodo')
            )
            totales = qs.aggregate(
                total_ventas=Sum('total'),
                cantidad_ventas=Count('id'),
            )
        except Exception:
            resumen = []
            totales = {'total_ventas': 0, 'cantidad_ventas': 0}

        return Response({'periodo': periodo, 'totales': totales, 'detalle': resumen})


class ReporteStockActualView(APIView):
    """GET /api/reportes/stock/?bajo_stock=true"""

    def get(self, request):
        solo_bajo_stock = request.query_params.get('bajo_stock', '').lower() in ('true', '1')

        try:
            # Usamos double-underscore en .values() para evitar conflicto con variante_id (campo FK)
            qs = list(
                Stock.objects
                .select_related('variante__producto')
                .values(
                    'id',
                    'cantidad',
                    'stock_minimo',
                    'variante__id',                    # ← sin alias para evitar conflicto
                    'variante__nombre',
                    'variante__sku',
                    'variante__precio_venta',
                    'variante__producto__nombre',
                )
                .order_by('variante__producto__nombre', 'variante__nombre')
            )

            # Filtro bajo stock en Python (evita F() en queryset que puede conflictuar)
            resultado = []
            for item in qs:
                bajo = item['cantidad'] <= item['stock_minimo']
                if solo_bajo_stock and not bajo:
                    continue
                resultado.append({
                    'variante_id':     item['variante__id'],
                    'producto_nombre': item['variante__producto__nombre'],
                    'variante_nombre': item['variante__nombre'],
                    'sku':             item['variante__sku'],
                    'precio_venta':    item['variante__precio_venta'],
                    'cantidad':        item['cantidad'],
                    'stock_minimo':    item['stock_minimo'],
                    'bajo_stock':      bajo,
                })

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'total_variantes': len(resultado), 'stock': resultado})


class ReporteProductosMasVendidosView(APIView):
    """GET /api/reportes/mas-vendidos/?limite=10&dias=30"""

    def get(self, request):
        limite = int(request.query_params.get('limite', 10))
        dias   = int(request.query_params.get('dias', 30))

        fecha_desde = timezone.now() - timedelta(days=dias)

        try:
            # Double-underscore en .values() sin alias para evitar el bug de Django 6
            raw = list(
                VentaDetalle.objects
                .filter(venta__fecha__gte=fecha_desde, venta__estado='COMPLETADA')
                .values(
                    'variante__id',
                    'variante__nombre',
                    'variante__producto__nombre',
                    'variante__sku',
                )
                .annotate(
                    total_unidades=Sum('cantidad'),
                    total_ingresos=Sum('subtotal'),
                    cantidad_ventas=Count('venta', distinct=True),
                )
                .order_by('-total_unidades')[:limite]
            )

            productos = [
                {
                    'variante_id':     item['variante__id'],
                    'variante_nombre': item['variante__nombre'],
                    'producto_nombre': item['variante__producto__nombre'],
                    'sku':             item['variante__sku'],
                    'total_unidades':  item['total_unidades'],
                    'total_ingresos':  str(item['total_ingresos'] or 0),
                    'cantidad_ventas': item['cantidad_ventas'],
                }
                for item in raw
            ]
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'periodo_dias': dias,
            'fecha_desde':  str(fecha_desde.date()),
            'productos':    productos,
        })


class ResumenDashboardView(APIView):
    """GET /api/reportes/dashboard/"""

    def get(self, request):
        hoy        = timezone.localdate()
        inicio_mes = hoy.replace(day=1)
        hace_7     = hoy - timedelta(days=7)

        try:
            ventas_hoy    = Venta.objects.filter(fecha__date=hoy, estado='COMPLETADA')
            ventas_semana = Venta.objects.filter(fecha__date__gte=hace_7, estado='COMPLETADA')
            ventas_mes    = Venta.objects.filter(fecha__date__gte=inicio_mes, estado='COMPLETADA')

            total_hoy    = ventas_hoy.aggregate(t=Sum('total'))['t'] or 0
            total_semana = ventas_semana.aggregate(t=Sum('total'))['t'] or 0
            total_mes    = ventas_mes.aggregate(t=Sum('total'))['t'] or 0

            # Usamos Python para el cálculo bajo_stock en vez de F() para evitar el conflicto
            todos_stocks  = list(Stock.objects.values('cantidad', 'stock_minimo'))
            bajo_stock    = sum(1 for s in todos_stocks if s['cantidad'] <= s['stock_minimo'])
            total_variantes = ProductoVariante.objects.filter(activo=True).count()

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'ventas_hoy':    {'cantidad': ventas_hoy.count(), 'total': total_hoy},
            'ventas_semana': {'total': total_semana},
            'ventas_mes':    {'total': total_mes},
            'inventario':    {
                'total_variantes_activas': total_variantes,
                'variantes_bajo_stock':    bajo_stock,
            },
            'fecha': str(hoy),
        })


class ExportarVentasExcelView(APIView):
    """GET /api/reportes/exportar/ventas/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD"""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin    = request.query_params.get('fecha_fin')

        qs = Venta.objects.prefetch_related('detalles__variante__producto').order_by('-fecha')
        if fecha_inicio:
            qs = qs.filter(fecha__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha__date__lte=fecha_fin)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ventas'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F46E5')

        headers = ['Comprobante','Fecha','Estado','M�todo Pago','Subtotal','Descuento','Total','Observaciones']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, venta in enumerate(qs, 2):
            ws.cell(row=row_num, column=1, value=venta.numero_comprobante)
            ws.cell(row=row_num, column=2, value=venta.fecha.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row_num, column=3, value=venta.get_estado_display())
            ws.cell(row=row_num, column=4, value=venta.get_metodo_pago_display())
            ws.cell(row=row_num, column=5, value=int(venta.subtotal))
            ws.cell(row=row_num, column=6, value=int(venta.descuento))
            ws.cell(row=row_num, column=7, value=int(venta.total))
            ws.cell(row=row_num, column=8, value=venta.observaciones)

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fname = f'ventas_{fecha_inicio or "todas"}_{fecha_fin or ""}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response


class ExportarStockExcelView(APIView):
    """GET /api/reportes/exportar/stock/ ? Excel con stock actual"""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        stocks = (
            Stock.objects
            .select_related('variante__producto')
            .order_by('variante__producto__nombre')
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='059669')

        headers = ['Producto','Variante','SKU','Precio Venta','Precio Costo','Stock Actual','Stock M�nimo','Estado']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, s in enumerate(stocks, 2):
            v = s.variante
            ws.cell(row=row_num, column=1, value=v.producto.nombre)
            ws.cell(row=row_num, column=2, value=v.nombre)
            ws.cell(row=row_num, column=3, value=v.sku)
            ws.cell(row=row_num, column=4, value=int(v.precio_venta))
            ws.cell(row=row_num, column=5, value=int(v.precio_costo))
            ws.cell(row=row_num, column=6, value=s.cantidad)
            ws.cell(row=row_num, column=7, value=s.stock_minimo)
            ws.cell(row=row_num, column=8, value='?? BAJO' if s.bajo_stock else 'OK')

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="stock_actual.xlsx"'
        wb.save(response)
        return response
